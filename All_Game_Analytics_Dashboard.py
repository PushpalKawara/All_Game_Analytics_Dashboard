import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import tempfile

# Initialize Streamlit app
st.set_page_config(page_title="Game Analytics Tool", layout="wide")
st.title("🎮 Game Level Data Analyzer")

# ======================== DATA PROCESSING FUNCTIONS ========================
def clean_level(level):
    """Extract numeric value from LEVEL column"""
    if pd.isna(level):
        return 0
    return int(re.sub(r'\D', '', str(level)))

def process_files(start_df, complete_df):
    """Process and merge the two dataframes with flexible column name handling."""
    def get_column(df, possible_names):
        """Return the first matching column name from the dataframe."""
        for col in df.columns:
            if col.strip().lower() in [name.lower() for name in possible_names]:
                return col
        return None

    # Flexible column matching
    level_col = get_column(start_df, ['LEVEL', 'TOTALLEVELS', 'STAGE'])
    game_col = get_column(start_df, ['GAME_ID', 'CATEGORY', 'Game_name' , 'MISSION'])
    diff_col = get_column(start_df, ['DIFFICULTY', 'mode'])

    playtime_col = get_column(complete_df, ['PLAY_TIME_AVG', 'PLAYTIME', 'PLAYTIME_AVG', 'playtime_avg'])
    hint_col = get_column(complete_df, ['HINT_USED_SUM', 'HINT_USED', 'HINT'])
    skipped_col = get_column(complete_df, ['SKIPPED_SUM', 'SKIPPED', 'SKIP'])
    attempts_col = get_column(complete_df, ['ATTEMPTS_SUM', 'ATTEMPTS', 'TRY_COUNT'])
    retry_col = get_column(complete_df, ['RETRY_SUM', 'RETRY'])

    # Clean LEVELs
    for df in [start_df, complete_df]:
        if level_col:
            df[level_col] = df[level_col].apply(clean_level)
            df.sort_values(level_col, inplace=True)

    # Rename required columns
    rename_dict_start = {'USERS': 'Start Users'}
    if level_col:
        rename_dict_start[level_col] = 'LEVEL'
    if game_col:
        rename_dict_start[game_col] = 'GAME_ID'
    if diff_col:
        rename_dict_start[diff_col] = 'DIFFICULTY'
    start_df.rename(columns=rename_dict_start, inplace=True)

    rename_dict_complete = {}
    if level_col:
        rename_dict_complete[level_col] = 'LEVEL'
    if game_col:
        rename_dict_complete[game_col] = 'GAME_ID'
    if diff_col:
        rename_dict_complete[diff_col] = 'DIFFICULTY'
    if playtime_col:
        rename_dict_complete[playtime_col] = 'PLAY_TIME_AVG'
    if hint_col:
        rename_dict_complete[hint_col] = 'HINT_USED_SUM'
    if skipped_col:
        rename_dict_complete[skipped_col] = 'SKIPPED_SUM'
    if attempts_col:
        rename_dict_complete[attempts_col] = 'ATTEMPTS_SUM'
    rename_dict_complete['USERS'] = 'Complete Users'
    complete_df.rename(columns=rename_dict_complete, inplace=True)

    # Merge
    merge_cols = []
    if 'GAME_ID' in start_df.columns:
        merge_cols.append('GAME_ID')
    if 'DIFFICULTY' in start_df.columns:
        merge_cols.append('DIFFICULTY')
    if 'LEVEL' in start_df.columns:
        merge_cols.append('LEVEL')
    merged = pd.merge(start_df, complete_df, on=merge_cols, how='outer', suffixes=('_start', '_complete'))

    # Build dynamic column list
    keep_cols = []
    if 'GAME_ID' in merged.columns:
        keep_cols.append('GAME_ID')
    if 'DIFFICULTY' in merged.columns:
        keep_cols.append('DIFFICULTY')
    if 'LEVEL' in merged.columns:
        keep_cols.append('LEVEL')
    keep_cols.extend(['Start Users', 'Complete Users'])
    if playtime_col and 'PLAY_TIME_AVG' in merged.columns:
        keep_cols.append('PLAY_TIME_AVG')
    if hint_col and 'HINT_USED_SUM' in merged.columns:
        keep_cols.append('HINT_USED_SUM')
    if skipped_col and 'SKIPPED_SUM' in merged.columns:
        keep_cols.append('SKIPPED_SUM')
    if attempts_col and 'ATTEMPTS_SUM' in merged.columns:
        keep_cols.append('ATTEMPTS_SUM')

    merged = merged[[col for col in keep_cols if col in merged.columns]]

    # Calculate drops and retention
    if 'Start Users' in merged.columns and 'Complete Users' in merged.columns:
        merged['Game Play Drop'] = ((merged['Start Users'] - merged['Complete Users']) / merged['Start Users'].replace(0, np.nan)) * 100
        merged['Popup Drop'] = ((merged['Complete Users'] - merged['Start Users'].shift(-1)) / merged['Complete Users'].replace(0, np.nan)) * 100
    else:
        merged['Game Play Drop'] = 0
        merged['Popup Drop'] = 0

    def calculate_retention(group):
        """Calculate retention using level 1/2 start users as base"""
        if 'Start Users' not in group.columns:
            group['Retention %'] = 0
            return group
        base_users = group[group['LEVEL'].isin([1, 2])]['Start Users'].max()
        if base_users == 0 or pd.isnull(base_users):
            base_users = group['Start Users'].max()
        group['Retention %'] = (group['Start Users'] / base_users) * 100
        return group

    # Determine grouping columns
    group_cols = []
    if 'GAME_ID' in merged.columns:
        group_cols.append('GAME_ID')
    if 'DIFFICULTY' in merged.columns:
        group_cols.append('DIFFICULTY')
    if not group_cols:
        merged['All Data'] = 'All Data'
        group_cols = ['All Data']
    merged = merged.groupby(group_cols, group_keys=False).apply(calculate_retention)

    # Fill NaN values
    fill_cols = ['Start Users', 'Complete Users']
    key_columns = ['PLAY_TIME_AVG', 'HINT_USED_SUM', 'SKIPPED_SUM', 'ATTEMPTS_SUM', 'RETRY_SUM']
    for col in key_columns:
        if col in merged.columns:
            fill_cols.append(col)
    merged.fillna({col: 0 for col in fill_cols}, inplace=True)

    if 'Game Play Drop' in merged.columns and 'Popup Drop' in merged.columns:
        merged['Total Level Drop'] = merged['Game Play Drop'] + merged['Popup Drop']
    else:
        merged['Total Level Drop'] = 0

    return merged


# ======================== CHART GENERATION ========================
def create_charts(df, game_name):
    """Generate enhanced matplotlib charts (levels 1–100 only)"""
    charts = {}

    # Filter up to level 100 only
    df_100 = df[df['LEVEL'] <= 100]

    # Custom x tick labels
    xtick_labels = []
    for val in np.arange(1, 101, 1):
        if val % 5 == 0:
            xtick_labels.append(f"$\\bf{{{val}}}$")  # Bold using LaTeX
        else:
            xtick_labels.append(str(val))


    # ========== RETENTION CHART ==========
    fig1, ax1 = plt.subplots(figsize=(15, 5))
    if 'Retention %' in df_100.columns and not df_100['Retention %'].dropna().empty:
        ax1.plot(df_100['LEVEL'], df_100['Retention %'],
                 linestyle='-', color='#F57C00', linewidth=2, label='Retention')

        ax1.set_xlim(1, 100)
        ax1.set_ylim(0, 110)
        ax1.set_xticks(np.arange(1, 101, 1))
        ax1.set_yticks(np.arange(0, 111, 5))
        ax1.set_xticklabels(xtick_labels, fontsize=4)
        ax1.tick_params(axis='x', labelsize=6)
        ax1.grid(True, linestyle='--', linewidth=0.5)

        ax1.set_xlabel("Level", labelpad=15)
        ax1.set_ylabel("% Of Users", labelpad=15)
        ax1.set_title(f"{game_name} | Retention Chart (Levels 1–100)",
                      fontsize=12, fontweight='bold')
        ax1.legend(loc='lower left', fontsize=8)

        for x, y in zip(df_100['LEVEL'], df_100['Retention %']):
            if not np.isnan(y):
                ax1.text(x, -5, f"{int(y)}", ha='center', va='top', fontsize=5)

    charts['retention'] = fig1

    # ========== TOTAL DROP CHART ==========
    fig2, ax2 = plt.subplots(figsize=(15, 5))
    if 'Total Level Drop' in df_100.columns and not df_100['Total Level Drop'].dropna().empty:
        bars = ax2.bar(df_100['LEVEL'], df_100['Total Level Drop'],
                       color='#EF5350', label='Drop Rate')

        drop_max = df_100['Total Level Drop'].max()
        drop_max = drop_max if not pd.isna(drop_max) else 0
        ymax = max(drop_max, 10) + 10

        ax2.set_xlim(1, 100)
        ax2.set_ylim(0, ymax)
        ax2.set_xticks(np.arange(1, 101, 1))
        ax2.set_yticks(np.arange(0, ymax + 1, 5))
        ax2.set_xticklabels(xtick_labels, fontsize=4)
        ax2.tick_params(axis='x', labelsize=6)
        ax2.grid(True, linestyle='--', linewidth=0.5)

        ax2.set_xlabel("Level")
        ax2.set_ylabel("% Of Users Drop")
        ax2.set_title(f"{game_name} | Total Drop Chart (Levels 1–100)",
                      fontsize=12, fontweight='bold')
        ax2.legend(loc='upper right', fontsize=8)

        for bar in bars:
            x = bar.get_x() + bar.get_width() / 2
            y = bar.get_height()
            ax2.text(x, -2, f"{y:.0f}", ha='center', va='top', fontsize=5)

    charts['total_drop'] = fig2

    # ========== COMBO DROP CHART ==========
    fig3, ax3 = plt.subplots(figsize=(15, 5))
    if ('Game Play Drop' in df_100.columns and
        'Popup Drop' in df_100.columns and
        not df_100['Game Play Drop'].dropna().empty and
        not df_100['Popup Drop'].dropna().empty):

        width = 0.4
        x = df_100['LEVEL']
        ax3.bar(x - width/2, df_100['Popup Drop'], width,
                color='#42A5F5', label='Popup Drop')
        ax3.bar(x + width/2, df_100['Game Play Drop'], width,
                color='#66BB6A', label='Game Play Drop')

        gpd_max = df_100['Game Play Drop'].max()
        pd_max = df_100['Popup Drop'].max()
        gpd_max = gpd_max if not pd.isna(gpd_max) else 0
        pd_max = pd_max if not pd.isna(pd_max) else 0
        max_drop = max(gpd_max, pd_max, 10) + 10

        ax3.set_xlim(1, 100)
        ax3.set_ylim(0, max_drop)
        ax3.set_xticks(np.arange(1, 101, 1))
        ax3.set_yticks(np.arange(0, max_drop + 1, 5))
        ax3.set_xticklabels(xtick_labels, fontsize=4)
        ax3.tick_params(axis='x', labelsize=6)
        ax3.grid(True, linestyle='--', linewidth=0.5)

        ax3.set_xlabel("Level")
        ax3.set_ylabel("% Of Users Dropped")
        ax3.set_title(f"{game_name} | Game Play & Popup Drop (Levels 1–100)",
                      fontsize=10, fontweight='bold')
        ax3.legend(loc='upper right', fontsize=6)

    charts['combined_drop'] = fig3

    return charts


# ======================== EXCEL GENERATION ========================
def generate_excel(processed_data):
    """Create Excel workbook with formatted sheets"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    main_sheet = wb.create_sheet("MAIN_TAB")
    main_headers = ["Index", "Sheet Name", "Link to Sheet", "LEVEL_Start", "Start Users",
                    "LEVEL_End", "USERS_END", "Game Play Drop Count", "Popup Drop Count",
                    "Total Level Drop Count"]
    main_sheet.append(main_headers)

    # Format main headers
    for col in main_sheet[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill("solid", fgColor="4F81BD")

    # Create sheets for each group
    for idx, (game_key, df) in enumerate(processed_data.items(), start=1):
        sheet_name = str(game_key)[:31]
        ws = wb.create_sheet(sheet_name)

        headers = ["=HYPERLINK(\"#MAIN_TAB!A1\", \"Back to Main\")", "Start Users", "Complete Users",
                   "Game Play Drop", "Popup Drop", "Total Level Drop", "Retention %",
                   "PLAY_TIME_AVG", "HINT_USED_SUM", "SKIPPED_SUM", "ATTEMPTS_SUM", "RETRY_SUM"]
        ws.append(headers)
        ws['A1'].font = Font(color="0000FF", underline="single", bold=True, size=14)
        ws['A1'].fill = PatternFill("solid", fgColor="FFFF00")
        ws.column_dimensions['A'].width = 25

        # Add data rows
        for _, row in df.iterrows():
            row_values = [
                row.get('LEVEL', 0),
                row.get('Start Users', 0),
                row.get('Complete Users', 0),
                round(row.get('Game Play Drop', 0), 2),
                round(row.get('Popup Drop', 0), 2),
                round(row.get('Total Level Drop', 0), 2),
                round(row.get('Retention %', 0), 2),
                round(row.get('PLAY_TIME_AVG', 0), 2),
                round(row.get('HINT_USED_SUM', 0), 2),
                round(row.get('SKIPPED_SUM', 0), 2),
                round(row.get('ATTEMPTS_SUM', 0), 2),
                round(row.get('RETRY_SUM', 0), 2),   
            ]
            ws.append(row_values)

        # Apply formatting and charts
        apply_sheet_formatting(ws)
        apply_conditional_formatting(ws, df.shape[0])
        charts = create_charts(df, sheet_name)
        add_charts_to_excel(ws, charts)

        # Update MAIN_TAB
        main_row = [
            idx, sheet_name,
            f'=HYPERLINK("#{sheet_name}!A1", "View")',
            df.get('LEVEL', 0).min(),
            df.get('Start Users', 0).max(),
            df.get('LEVEL', 0).max(),
            df.get('Complete Users', 0).iloc[-1] if not df.empty else 0,
            sum(df.get('Game Play Drop', 0) >= 3),
            sum(df.get('Popup Drop', 0) >= 3),
            sum(df.get('Total Level Drop', 0) >= 3)
            
        ]
        main_sheet.append(main_row)

         # Apply formatting to all cells in main sheet
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 1:  # Keep header formatting
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")


    column_widths = [8, 25, 20, 18, 20, 12, 15, 12, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        main_sheet.column_dimensions[get_column_letter(i)].width = width

    return wb

# ======================== HELPER FUNCTIONS ========================
def apply_sheet_formatting(sheet):
    """Apply consistent formatting to sheets"""
    sheet.freeze_panes = 'A2'
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
 #  Special formatting for A1 only in game sheets (not main tab)
    if sheet.title != "MAIN_TAB":
        a1_cell = sheet['A1']
        a1_cell.font = Font(color="0000FF", underline="single", bold=True, size=11)
        a1_cell.fill = PatternFill("solid", fgColor="FFFF00")
        sheet.column_dimensions['A'].width = 14
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-fit columns for other columns
    for col in sheet.columns:
        if col[0].column == 1 and sheet.title != "MAIN_TAB":  # Skip column A for game sheets
            continue
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2



def apply_conditional_formatting(sheet, num_rows):
    """Apply color scale formatting to drop columns"""
    for row in sheet.iter_rows(min_row=2, max_row=num_rows+1):
        for cell in row:
            if cell.column_letter in ['D', 'E', 'F'] and isinstance(cell.value, (int, float)):
                if cell.value >= 10:
                    cell.fill = PatternFill(start_color='990000', end_color='990000', fill_type='solid')
                    cell.font = Font(color="FFFFFF")
                elif cell.value >= 7:
                    cell.fill = PatternFill(start_color='CC3333', end_color='CC3333', fill_type='solid')
                    cell.font = Font(color="FFFFFF")
                elif cell.value >= 3:
                    cell.fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
                    cell.font = Font(color="FFFFFF")
                elif cell.value < 0:
                    # Do not apply any fill or font color
                    pass
            cell.alignment = Alignment(horizontal='center', vertical='center')

def add_charts_to_excel(worksheet, charts):
    """Add charts to Excel worksheet"""
    img_positions = {'retention': 'M2', 'total_drop': 'M52', 'combined_drop': 'M98'}
    for chart_type, pos in img_positions.items():
        if chart_type in charts:
            img_data = BytesIO()
            charts[chart_type].savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            img = OpenpyxlImage(img_data)
            worksheet.add_image(img, pos)
            plt.close(charts[chart_type])

# ======================== STREAMLIT UI ========================
def main():
    st.sidebar.header("Upload Files")
    start_file = st.sidebar.file_uploader("LEVEL_START.csv", type="csv")
    complete_file = st.sidebar.file_uploader("LEVEL_COMPLETE.csv", type="csv")

    if start_file and complete_file:
        with st.spinner("Processing data..."):
            try:
                start_df = pd.read_csv(start_file)
                complete_df = pd.read_csv(complete_file)
                merged = process_files(start_df, complete_df)

                # Determine grouping columns
                group_cols = []
                if 'GAME_ID' in merged.columns:
                    group_cols.append('GAME_ID')
                if 'DIFFICULTY' in merged.columns:
                    group_cols.append('DIFFICULTY')
                if not group_cols:
                    if 'All Data' not in merged.columns:
                        merged['All Data'] = 'All Data'
                    group_cols = ['All Data']

                processed_data = {}
                for group_key, group_df in merged.groupby(group_cols):
                    key = '_'.join(map(str, group_key)) if isinstance(group_key, tuple) else str(group_key)
                    processed_data[key] = group_df

                # Generate Excel
                wb = generate_excel(processed_data)
                with tempfile.NamedTemporaryFile(delete=False) as tmp:
                    wb.save(tmp.name)
                    with open(tmp.name, "rb") as f:
                        excel_bytes = f.read()

                st.success("Processing complete!")
                st.download_button(
                    label="📥 Download Consolidated Report",
                    data=excel_bytes,
                    file_name="Game_Analytics_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                with st.expander("Preview Processed Data"):
                    st.dataframe(merged.head(20))

            except Exception as e:
                st.error(f"Error processing files: {str(e)}")

if __name__ == "__main__":
    main()
